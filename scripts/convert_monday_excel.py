#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Monday.com Excel Export to Nigiwai PM Converter
Converts Monday.com exported .xlsx files to JSON structure and MySQL SQL Seed files.
"""

import os
import sys
import json
import argparse
import openpyxl
from datetime import datetime, date

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def detect_column_type(title, sample_values):
    title_lower = str(title).lower().strip()
    if any(k in title_lower for k in ['status', 'priority', 'state']):
        return 'status'
    if any(k in title_lower for k in ['date', 'timeline', 'opening', 'start', 'end', 'updated', 'due']):
        return 'date'
    if any(k in title_lower for k in ['progress', '%', 'duration', 'formula', 'overall']):
        return 'progress' if '%' in title_lower or 'progress' in title_lower else 'number'
    if any(k in title_lower for k in ['owner', 'people', 'assing', 'assign', 'contacts', 'user']):
        return 'people'
    return 'text'

def clean_cell_value(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d %H:%M:%S' if isinstance(val, datetime) else '%Y-%m-%d')
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    return s if s else None

def escape_sql_val(val):
    if val is None:
        return 'NULL'
    if isinstance(val, (int, float)):
        return str(val)
    # Double escape backslashes for MySQL string literals and escape single quotes
    s = str(val).replace('\\', '\\\\').replace("'", "''")
    return f"'{s}'"

def escape_sql_json(obj):
    if obj is None:
        return "'{}'"
    json_text = json.dumps(obj, ensure_ascii=False)
    # In MySQL single-quoted strings, JSON escape sequences like \" must be written as \\\"
    s = json_text.replace('\\', '\\\\').replace("'", "''")
    return f"'{s}'"

def parse_monday_excel(excel_path):
    print(f"[+] Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 1. Main Sheet
    main_sheet_name = wb.sheetnames[0]
    sheet = wb[main_sheet_name]
    
    board_name = clean_cell_value(sheet.cell(row=1, column=1).value) or "Imported Board"
    board_desc = clean_cell_value(sheet.cell(row=2, column=1).value) or ""
    
    groups = []
    current_group = None
    current_item = None
    
    main_headers = []
    sub_headers = []
    
    main_col_map = {} # col_title -> col_id
    sub_col_map = {}  # col_title -> col_id
    
    all_main_columns = []
    all_sub_columns = []
    
    group_idx = 0
    main_item_idx = 0
    subitem_idx = 0
    
    colors_palette = ['#579BFC', '#00C875', '#E2445C', '#A25DDC', '#FDAB3D', '#FF642E', '#0086C0', '#784BD1']
    
    for r in range(1, sheet.max_row + 1):
        c1 = sheet.cell(row=r, column=1).value
        c2 = sheet.cell(row=r, column=2).value
        
        if r in [1, 2, 3]:
            continue
            
        # Group Header Row
        if c1 is not None and c2 is None:
            next_c1 = sheet.cell(row=r+1, column=1).value if r+1 <= sheet.max_row else None
            if next_c1 == 'Name':
                group_idx += 1
                group_title = str(c1).strip()
                color = colors_palette[(group_idx - 1) % len(colors_palette)]
                current_group = {
                    'id': group_idx,
                    'title': group_title,
                    'color': color,
                    'position': float(group_idx),
                    'items': []
                }
                groups.append(current_group)
                current_item = None
                continue
                
        # Main Items Header Row
        if c1 == 'Name':
            headers = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val:
                    val_str = str(val).strip()
                    headers.append(val_str)
                    if val_str not in main_col_map and val_str != 'Name':
                        col_id = f"col_{len(main_col_map)+1}"
                        col_type = detect_column_type(val_str, [])
                        main_col_map[val_str] = col_id
                        all_main_columns.append({
                            'id': col_id,
                            'title': val_str,
                            'type': col_type,
                            'is_subitem': 0,
                            'position': float(len(all_main_columns) + 1),
                            'settings': {}
                        })
            main_headers = headers
            continue
            
        # Subitems Header Row
        if c1 == 'Subitems':
            headers = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val:
                    val_str = str(val).strip()
                    headers.append(val_str)
                    if val_str not in sub_col_map and val_str not in ['Subitems', 'Name']:
                        col_id = f"sub_col_{len(sub_col_map)+1}"
                        col_type = detect_column_type(val_str, [])
                        sub_col_map[val_str] = col_id
                        all_sub_columns.append({
                            'id': col_id,
                            'title': val_str,
                            'type': col_type,
                            'is_subitem': 1,
                            'position': float(len(all_sub_columns) + 1),
                            'settings': {}
                        })
            sub_headers = headers
            continue
            
        # Main Item Row
        if c1 is not None and c1 not in [board_name, 'Name', 'Subitems'] and not str(c1).startswith('Track'):
            main_item_idx += 1
            item_name = str(c1).strip()
            col_vals = {}
            monday_id = None
            
            for c in range(1, len(main_headers) + 1):
                h = main_headers[c-1]
                val = clean_cell_value(sheet.cell(row=r, column=c).value)
                if 'Item ID' in h and val:
                    monday_id = str(val)
                elif h in main_col_map and val is not None:
                    col_vals[main_col_map[h]] = val
                    
            current_item = {
                'id': main_item_idx,
                'monday_item_id': monday_id,
                'name': item_name,
                'column_values': col_vals,
                'position': float(len(current_group['items']) + 1) if current_group else 1.0,
                'subitems': []
            }
            if current_group is not None:
                current_group['items'].append(current_item)
            continue
            
        # Subitem Row
        if c1 is None and c2 is not None and current_item is not None:
            subitem_idx += 1
            sub_name = str(c2).strip()
            sub_col_vals = {}
            monday_sub_id = None
            
            for c in range(1, len(sub_headers) + 1):
                h = sub_headers[c-1]
                val = clean_cell_value(sheet.cell(row=r, column=c).value)
                if 'Item ID' in h and val:
                    monday_sub_id = str(val)
                elif h in sub_col_map and val is not None:
                    sub_col_vals[sub_col_map[h]] = val
                    
            sub_item = {
                'id': subitem_idx,
                'monday_item_id': monday_sub_id,
                'name': sub_name,
                'column_values': sub_col_vals,
                'position': float(len(current_item['subitems']) + 1)
            }
            current_item['subitems'].append(sub_item)
            
    # 2. Updates Sheet (if present)
    updates = []
    if 'updates' in wb.sheetnames:
        u_sheet = wb['updates']
        for r in range(3, u_sheet.max_row + 1):
            item_id = clean_cell_value(u_sheet.cell(row=r, column=1).value)
            if item_id is None:
                continue
            item_name = clean_cell_value(u_sheet.cell(row=r, column=2).value)
            user_name = clean_cell_value(u_sheet.cell(row=r, column=5).value) or "Unknown User"
            created_at = clean_cell_value(u_sheet.cell(row=r, column=6).value) or ""
            content = clean_cell_value(u_sheet.cell(row=r, column=7).value) or ""
            likes = clean_cell_value(u_sheet.cell(row=r, column=8).value) or 0
            post_id = clean_cell_value(u_sheet.cell(row=r, column=10).value)
            
            updates.append({
                'monday_item_id': str(item_id),
                'item_name': item_name,
                'user_name': str(user_name),
                'created_at': str(created_at),
                'content': str(content),
                'likes_count': int(likes) if isinstance(likes, (int, float)) else 0,
                'monday_post_id': str(post_id) if post_id else None
            })

    # Map updates to items and subitems
    updates_by_monday_id = {}
    for u in updates:
        mid = str(u['monday_item_id'])
        if mid not in updates_by_monday_id:
            updates_by_monday_id[mid] = []
        updates_by_monday_id[mid].append(u)

    for g in groups:
        for item in g['items']:
            mid = str(item.get('monday_item_id') or '')
            item_ups = updates_by_monday_id.get(mid, [])
            item['update_count'] = len(item_ups)
            item['updates'] = item_ups
            for sub in item.get('subitems', []):
                sub_mid = str(sub.get('monday_item_id') or '')
                sub_ups = updates_by_monday_id.get(sub_mid, [])
                sub['update_count'] = len(sub_ups)
                sub['updates'] = sub_ups
            
    board_data = {
        'board': {
            'name': board_name,
            'description': board_desc
        },
        'columns': all_main_columns + all_sub_columns,
        'groups': groups,
        'updates': updates,
        'stats': {
            'total_groups': len(groups),
            'total_main_items': main_item_idx,
            'total_subitems': subitem_idx,
            'total_updates': len(updates)
        }
    }
    
    return board_data

def generate_sql_dump(board_data, sql_path):
    print(f"[+] Generating SQL dump: {sql_path}")
    with open(sql_path, 'w', encoding='utf-8') as f:
        f.write("-- Nigiwai PM Generated Seed Data\n")
        f.write("SET NAMES utf8mb4;\n")
        f.write("SET FOREIGN_KEY_CHECKS = 0;\n\n")
        
        # Clean existing data to avoid duplicate key errors on re-import
        f.write("DELETE FROM `item_updates`;\n")
        f.write("DELETE FROM `items`;\n")
        f.write("DELETE FROM `board_columns`;\n")
        f.write("DELETE FROM `board_groups`;\n")
        f.write("DELETE FROM `boards`;\n\n")
        
        # 1. Board
        b_name = escape_sql_val(board_data['board']['name'])
        b_desc = escape_sql_val(board_data['board']['description'])
        f.write(f"REPLACE INTO `boards` (`id`, `name`, `description`) VALUES (1, {b_name}, {b_desc});\n\n")
        
        # 2. Columns
        for col in board_data['columns']:
            c_id = col['id']
            c_title = escape_sql_val(col['title'])
            c_type = col['type']
            c_is_sub = col['is_subitem']
            c_pos = col['position']
            c_settings = escape_sql_json(col['settings'])
            f.write(f"REPLACE INTO `board_columns` (`id`, `board_id`, `title`, `type`, `is_subitem`, `settings`, `position`) VALUES ('{c_id}', 1, {c_title}, '{c_type}', {c_is_sub}, {c_settings}, {c_pos});\n")
        f.write("\n")
        
        # 3. Groups & Items
        item_id_counter = 1
        item_mapping = {} # monday_item_id -> database_item_id
        
        for g in board_data['groups']:
            g_id = g['id']
            g_title = escape_sql_val(g['title'])
            g_color = escape_sql_val(g['color'])
            g_pos = g['position']
            f.write(f"REPLACE INTO `board_groups` (`id`, `board_id`, `title`, `color`, `position`) VALUES ({g_id}, 1, {g_title}, {g_color}, {g_pos});\n")
            
            for item in g['items']:
                db_item_id = item_id_counter
                item_id_counter += 1
                
                m_id = escape_sql_val(item['monday_item_id'])
                if item['monday_item_id']:
                    item_mapping[item['monday_item_id']] = db_item_id
                    
                i_name = escape_sql_val(item['name'])
                i_vals = escape_sql_json(item['column_values'])
                i_pos = item['position']
                f.write(f"REPLACE INTO `items` (`id`, `monday_item_id`, `board_id`, `group_id`, `parent_id`, `name`, `column_values`, `position`) VALUES ({db_item_id}, {m_id}, 1, {g_id}, NULL, {i_name}, {i_vals}, {i_pos});\n")
                
                for sub in item['subitems']:
                    db_sub_id = item_id_counter
                    item_id_counter += 1
                    
                    sub_m_id = escape_sql_val(sub['monday_item_id'])
                    if sub['monday_item_id']:
                        item_mapping[sub['monday_item_id']] = db_sub_id
                        
                    s_name = escape_sql_val(sub['name'])
                    s_vals = escape_sql_json(sub['column_values'])
                    s_pos = sub['position']
                    f.write(f"REPLACE INTO `items` (`id`, `monday_item_id`, `board_id`, `group_id`, `parent_id`, `name`, `column_values`, `position`) VALUES ({db_sub_id}, {sub_m_id}, 1, {g_id}, {db_item_id}, {s_name}, {s_vals}, {s_pos});\n")
                    
        f.write("\n")
        
        # 4. Updates
        for u in board_data['updates']:
            m_item_id = u['monday_item_id']
            if m_item_id in item_mapping:
                target_db_id = item_mapping[m_item_id]
                u_user = escape_sql_val(u['user_name'])
                u_content = escape_sql_val(u['content'])
                u_likes = u['likes_count']
                u_post_id = escape_sql_val(u['monday_post_id'])
                f.write(f"REPLACE INTO `item_updates` (`item_id`, `monday_post_id`, `user_name`, `content`, `likes_count`) VALUES ({target_db_id}, {u_post_id}, {u_user}, {u_content}, {u_likes});\n")
                
        f.write("\nSET FOREIGN_KEY_CHECKS = 1;\n")
    print(f"[OK] SQL dump generated successfully: {sql_path}")

def main():
    parser = argparse.ArgumentParser(description="Convert Monday.com Excel export to Nigiwai PM format")
    parser.add_argument("--excel", default="Branch_Planing_2026_1787817666.xlsx", help="Input Monday Excel file (.xlsx)")
    parser.add_argument("--output-json", default="data/board_data.json", help="Output JSON file path")
    parser.add_argument("--output-sql", default="data/seed_data.sql", help="Output SQL Seed file path")
    args = parser.parse_args()
    
    data = parse_monday_excel(args.excel)
    
    os.makedirs(os.path.dirname(args.output_json), exist_ok=True)
    with open(args.output_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON output saved: {args.output_json}")
    
    if args.output_sql:
        generate_sql_dump(data, args.output_sql)
        
    print("\n" + "="*50)
    print("SUMMARY OF CONVERSION:")
    print(f"Board Name: {data['board']['name']}")
    print(f"Total Groups: {data['stats']['total_groups']}")
    print(f"Total Main Items: {data['stats']['total_main_items']}")
    print(f"Total Subitems: {data['stats']['total_subitems']}")
    print(f"Total Updates: {data['stats']['total_updates']}")
    print("="*50)

if __name__ == '__main__':
    main()
